import os
import io
import random
import string
from openpyxl import load_workbook, Workbook
from functools import wraps
from flask import Flask, request, send_file, render_template, make_response, redirect, session,flash,url_for
from flask_mail import Mail, Message
import requests
import pandas as pd
from pyDes import *
import modelos
import datetime
import numpy as np
import shutil
from io import BytesIO

dominio = "https://ds.efigas.com.co/" # CAMBIAR AL DOMINIO DE DESPLIEGUE
app = Flask(__name__)
app.secret_key = b'f83ea3ca1a60a8ae08ae73681c5565870666a30bce8dfdb6a4fe65b9508fdd0a' ## LLAVE PARA ENVOLVER LA APLICACIÓN TIENE QUE SER EN EL FORMATO HEXADECIMAL Y LUEGO SE PASA A CADENA DE BYTES QUE TERMINA SIENDO DEL FORMATO b''
key = b"\xaf\xf5q'Pg\xf7\xeeC\x9e\xde\xf4FM,\xe6`\xe7\xd8\x01\xa0T\xaam" ## LLAVE PARA ENVOLVER LA APLICACIÓN TIENE QUE SER EN EL FORMATO HEXADECIMAL Y LUEGO SE PASA A CADENA DE BYTES QUE TERMINA SIENDO DEL FORMATO b'' ESTA SERA LA LLAVE PARA CODIFICAR LAS CONTRASEÑAS

# Flask-Mail configuration
app.config['MAIL_SERVER'] = 'smtp.gmail.com'  # el servidor SMTP de preferencia asociado al servidor del correo es decir si es gmail o outlook etc 
app.config['MAIL_PORT'] = 587  # el puerto del servidor SMTP que se requiere y por el cual se envian los correos
app.config['MAIL_USERNAME'] = 'lz7910412@gmail.com'  # VARIABLE DE ENTORNO QUE REPRESENTA EL EMAIL POR DONDE SE VAN A ENVIAR LOS CORREOS ASOCIADOS
app.config['MAIL_PASSWORD'] = 'zexw bofv jgbc jdyf'  # VARIABLE DE ENTORNO QUE CONTIENE LA CONTRASEÑA DE APLICACIÓN NO DEL EMAIL, SINO LA CONTRASEÑA DE APLICACIÓN QUE SE CREA EN EL PORTAL DE GMAIL O MICROSOFT Y QUE PERMITE ENVIAR CORREOS DESDE EL EMAIL SUMINISTRADO
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_DEFAULT_SENDER'] = 'lz7910412@gmail.com' # Correo desde el cual se envian mensajes si el mensaje no especifica se puede dejar el mismo MAIL_USERNAME
ail = Mail(app)

#Redirige a la pagina de inicio si se carga una página interna sin una sesión iniciada
# def login_required(f):
#     @wraps(f)
#     def decorated_function(*args, **kwargs):
#         if session.get('email') is None or session.get('ingreso') is None:
#             return redirect('/Inicio')
#         return f(*args, **kwargs)
#     return decorated_function

#Valida credenciales
@app.route('/Ingresar', methods=['POST'])
def iniciar():
    ## OBTENEMOS EL USUARIO
    identification = request.form.get('identification')
    password = request.form.get('password')
    
    ### LLAMAMOS EL ENDPOINT PARA INGRESAR EL USUARIO
    object_ = {
        "identification":identification,
        "password":password
    }
    print("OBJETO: ",object_)
    ### HACEMOS LA QUERY AL SERVIDOR
    url = dominio+"rupture/login"#"https://mongorupture.efigasprojecthub.site/rupture/login"
    
    response = requests.post(url,json=object_)
    
    if(response.json()['status'] == 'Registro éxitoso'):
        resp = make_response(redirect("/Perfil"))
        resp.set_cookie('ingreso', 'true')
        session['ingreso'] = True
        resp.set_cookie('email', response.json()['email'])
        session['email'] = response.json()['email']
        resp.set_cookie('nombre1', response.json()['first_name'])
        resp.set_cookie('nombre2', response.json()['last_name'])
        resp.set_cookie('empresa', 'Efigas S.A')
        resp.set_cookie('tipo_user', 'worker')
        return resp
    else:
        ### MOSTRAMOS UNA ALERTA INDICANDO EL PROBLEMA
        flash(response.json()['status'], 'error')
        resp = make_response(redirect("/Inicio", code=307))
        return resp


@app.route('/Perfil')
def perfil():
    resp = make_response(render_template('principal.html'))
    resp.set_cookie('nuevo', '')
    return resp

@app.route('/')
def home():
    if request.cookies.get("ingreso") == "true":
        return redirect("/Perfil")
    return redirect("/Inicio")

#Cargar registro de usuario. Cierra sesión
@app.route('/Registrarse', methods=['GET'])
def nuevo_usuario():
    resp = make_response(render_template('crear_usuario.html', existe="false"))
    return resp

#Cargar registro de admin
@app.route('/RegistrarseAdmin', methods=['POST'])
def nuevo_admin():
    resp = make_response(render_template('crear_admin.html', existe="false"))
    return resp

#Cargar registro de usuario. Cierra sesión
@app.route('/OlvideMiContraseña', methods=['GET'])
def olvide_contraseña():
    resp = make_response(render_template('olvide_contra.html'))
    if request.cookies.get("ingreso") == "true":
        resp.set_cookie('ingreso', 'false')
        session.pop('ingreso', None)
        resp.set_cookie('email', '')
        session.pop('email', None)
        resp.set_cookie('nombre1', '')
        resp.set_cookie('nombre2', '')
        resp.set_cookie('empresa', '')
        resp.set_cookie('tipo_user', '')
    return resp

#Enviar correo de recuperación
@app.route('/OlvideMiContraseña', methods=['POST'])
def correoRecuperar():
    resp = make_response(redirect('/EnviarRecuperacion', code=307))
    email = request.form.get("email")
    df = pd.read_excel('usuarios.xlsx')
    #Verificar si existe usuario
    fila = df[(df['email'] == email)]
    if fila.shape[0] > 0:
        #Generar token
        token = ''.join(random.choices(string.ascii_uppercase + string.ascii_lowercase + string.digits, k=12))
        guardar_token(email, token, 2) #Token tipo 2: recuperar contraseña

        #Enviar correo
        verification_link = dominio + "/Recuperar?tk=" + token
        msg = Message('Restablece tu contraseña', recipients=[email])
        msg.body = f"Haz click en el enlace para restablecer tu contraseña: {verification_link}\n\nEl enlace expira en 30 minutos."
        mail.send(msg)
    return resp

@app.route('/EnviarRecuperacion', methods=['POST'])
def recuperarEnviado():
    resp = make_response(render_template('resta_enviado.html'))
    return resp

#Restablecer contraseña
@app.route('/Recuperar', methods=['GET'])
def recuperar_contraseña():
    token = request.args.get('tk')
    session.pop("restablecer", None)
    wb_tk = load_workbook("solicitudes.xlsx")
    ws_tk = wb_tk.active
    wb_us = load_workbook("usuarios.xlsx")
    ws_us = wb_us.active
    usuario = False
    fila = -1
    tiempo = datetime.datetime

    #Busca si existe el token para verificar usuario
    for row in ws_tk.iter_rows():
        if row[1].value == 2 and row[3].value == token:
            tiempo = datetime.datetime.now() - row[2].value
            email = row[0].value
            fila = row[0].row
            break
    #Busca si existe el usuario
    if fila > -1:
        for row in ws_us.iter_rows():
            if row[4].value == email:
                usuario = True
                break
    
    if (fila > -1):
        if usuario:
            if tiempo.total_seconds() < 1800:
                resp = make_response(render_template('recuperar.html',valido="true"))
                resp.set_cookie('restablecer', value=email)
            else:
                ws_tk.delete_rows(fila)
                wb_tk.save("solicitudes.xlsx")
                resp = make_response(render_template('recuperar.html',valido="false"))
        else:
            ws_tk.delete_rows(fila)
            wb_tk.save("solicitudes.xlsx")
            resp = make_response(render_template('recuperar.html',valido="false"))
    else:
        resp = make_response(render_template('recuperar.html',valido="false"))

    wb_tk.close()
    wb_us.close()
    return resp

#Guardar la nueva contraseña
@app.route('/Restablecido', methods=['POST'])
def guardar_contraseña():
    contraseña = request.form.get("contraseña")
    email = request.cookies.get("restablecer")
    wb_tk = load_workbook("solicitudes.xlsx")
    ws_tk = wb_tk.active
    wb_us = load_workbook("usuarios.xlsx")
    ws_us = wb_us.active
    usuario = False
    fila = -1

    #Busca la fila del token
    if email != None:
        for row in ws_tk.iter_rows():
            if row[0].value == email and row[1].value == 2:
                fila = row[0].row
                break
    #Busca la fila del usuario
    if fila > -1:
        ws_tk.delete_rows(fila)
        wb_tk.save("solicitudes.xlsx")
        for row in ws_us.iter_rows():
            if row[4].value == email:
                usuario = True
                pw = repr(triple_des(key).encrypt(contraseña, padmode=2))
                row[5].value = pw
                wb_us.save("usuarios.xlsx")
                break
    if usuario:
        resp = make_response(render_template('restablecido.html',exito="true"))
    else:
        resp = make_response(render_template('restablecido.html',exito="false"))
    wb_tk.close()
    wb_us.close()
    return resp

#Registrar nuevo usuario
@app.route('/Registrar', methods=['POST'])
def guardar_usuario():
    first_name = request.form.get("first_name")
    last_name = request.form.get("last_name")
    email = request.form.get("email")
    identification = request.form.get("identification")
    password = request.form.get("password")
    ## LOS DATOS LLEVAN BIEN CREAMOS EL USUARIO EN EL ESQUEMA DE LA BASE DE DATOS
    ## ARMAMOS EL BODY
    object_ = {
        "first_name":first_name,
        "last_name":last_name,
        "email":email,
        "identification":identification,
        "password":password,
        "state":True,
        "rol":"worker"
    }
    ### HACEMOS LA QUERY AL SERVIDOR
    url = dominio+"/rupture/createUser"

    response = requests.post(url,json=object_)

    if(response.json()['status'] == 'Usuario creado con éxito'):
        # flash('Usuario creado con éxito', 'success')
        return redirect(url_for('salir'))
    else:
        ### MOSTRAMOS UNA ALERTA INDICANDO EL PROBLEMA
        flash('Usuario con cédula ya registrada', 'error')
        return redirect(url_for('nuevo_usuario'))


    



@app.route('/Registrado', methods=['POST'])
def usuarioEnviado():
    resp = make_response(render_template("creado.html"))
    return resp

#Registrar nuevo administrador
@app.route('/RegistrarAdmin', methods=['POST'])
def guardar_admin():
    nombre1 = request.form.get("nombre1")
    nombre2 = request.form.get("nombre2")
    apellido1 = request.form.get("apellido1")
    apellido2 = request.form.get("apellido2")
    email = request.form.get("email")
    contraseña = request.form.get("contraseña1")
    df = pd.read_excel('usuarios.xlsx')
    
    fila = df[(df['email'] == email)]
    if fila.shape[0] > 0:
        resp = make_response(render_template('crear_admin.html',existe="true"))
    else:
        pw = triple_des(key).encrypt(contraseña, padmode=2)
        empresa = email.split('@')[1].split('.')[0]
        df = pd.DataFrame({'Nombre1':[nombre1], 'Nombre2':[nombre2], 'Apellido1':[apellido1], 'Apellido2':[apellido2], 'email':[email], 'contraseña':[pw], 'tipo':[1], 'empresa':[empresa], 'verificado':['no']})

        # Cargar el archivo CSV existente
        try:
            df_existente = pd.read_excel('usuarios.xlsx')
            df = pd.concat([df_existente, df], ignore_index=True)

            # Guardar los datos actualizados en el archivo excel
            df.to_excel('usuarios.xlsx', index=False)

            #Generar token
            token = ''.join(random.choices(string.ascii_uppercase + string.ascii_lowercase + string.digits, k=12))
            guardar_token(email, token, 1) #Token tipo 1: Verificacion de email

            #Enviar correo
            verification_link = dominio + "/Verificar?tk=" + token
            msg = Message('Verifica tu correo electrónico', recipients=[email])
            msg.body = f"Haz click en el enlace para verificar el correo electrónico: {verification_link}"
            mail.send(msg)
        except FileNotFoundError:
            pass
        if not os.path.exists("eventos/" + empresa + ".xlsx"):
            crear_tabla(empresa)
        resp = make_response(redirect('/AdminRegistrado', code=307))
    return resp

@app.route('/AdminRegistrado', methods=['POST'])
def adminEnviado():
    resp = make_response(render_template("creado_admin.html"))
    return resp

#Verificar correo de usuario
@app.route('/Verificar', methods=['GET'])
def verificar_usuario():
    token = request.args.get('tk')
    wb_tk = load_workbook("solicitudes.xlsx")
    ws_tk = wb_tk.active
    wb_us = load_workbook("usuarios.xlsx")
    ws_us = wb_us.active
    usuario = False
    borrar = -1

    #Busca si existe el token para verificar usuario
    for row in ws_tk.iter_rows():
        if row[1].value == 1 and row[3].value == token:
            email = row[0].value
            borrar = row[0].row
            break
    #Busca si existe el usuario
    if borrar > -1:
        for row in ws_us.iter_rows():
            if row[4].value == email:
                usuario = True
                break

    #si existe token y usuario, verifica y borra token
    if (borrar > -1):
        ws_tk.delete_rows(borrar)
        wb_tk.save("solicitudes.xlsx")
        if usuario:
            row[8].value = "si"
            wb_us.save("usuarios.xlsx")
        resp = make_response(render_template('verificar.html',valido="true"))
    else:
        resp = make_response(render_template('verificar.html',valido="false"))
    wb_tk.close()
    wb_us.close()
    return resp

# Salir de sesión
@app.route('/Inicio', methods=['POST', 'GET'])
def salir():
    if request.method == "POST":
        resp = make_response(render_template('index.html', mostrar = "true"))
    else:
        resp = make_response(render_template('index.html', mostrar = "false"))
    if request.cookies.get("ingreso") == "true":
        resp.set_cookie('ingreso', 'false')
        session.pop('ingreso', None)
        resp.set_cookie('email', '')
        session.pop('email', None)
        resp.set_cookie('nombre1', '')
        resp.set_cookie('nombre2', '')
        resp.set_cookie('empresa', '')
        resp.set_cookie('tipo_user', '')
    return resp

#Preparar nuevo evento
@app.route('/NuevoEvento', methods=['POST'])
def nuevo():
    resp = make_response(render_template('evento.html'))
    resp.set_cookie('guardado', 'false')
    return resp

@app.route('/Resultados', methods=['POST'])
def nuevoEvento():
    # Se obtienen todas las entradas

    orden = request.form.get('orden')
    ubicacion = request.form.get('ubicacion')
    presionTub = request.form.get('presion')
    presionUni = request.form.get('presionUni')
    subte = request.form.get('subte')
    equi = request.form.get('diameEqui')
    escape = request.form.get('escape')
    direccion = request.form.get("Flujo")
    forma = request.form.get('Forma')
    Fdiametro = request.form.get('DiameFuga')
    longitud = request.form.get('LongiFuga')
    Alto = request.form.get('Altofuga')
    FdiametroUni = request.form.get('DiameFugaUni')
    longitudUni = request.form.get('LongiFugaUni')
    AltoUni = request.form.get('AltofugaUni')
    Tlargo = request.form.get("DistTube")
    TlargoUni = request.form.get("DistTubeUni")
    Tlargo2 = request.form.get("DistTube2")
    TlargoUni2 = request.form.get("DistTubeUni2")
    Tdiametro = request.form.get('DiameTube')
    tiempoInicio = request.form.get('tiempoInicio')
    tiempoFin = request.form.get('tiempoFin')
    

    #Se castean los valores porque a numpy no le gusta usar cadenas
    
    presionTub = float(presionTub)
    Fdiametro = float(Fdiametro) if Fdiametro != "" else 0
    Tlargo = float(Tlargo) if Tlargo != "" else 0
    Tlargo2 = float(Tlargo2) if Tlargo2 != "" else 0
    longitud = float(longitud) if longitud != "" else 0
    #Alto = float(Alto) if Alto != "" else 0
    Tdiametro = float(Tdiametro)
    
    #Aplicar los diametros equivalentes
    if equi == 'on':
        Fuga_diame = convertir("in", "mm", modelos.diametro_equi(Tdiametro, escape))
    else:
        Fuga_diame = convertir(FdiametroUni, "mm", Fdiametro)
    #Convertir a mm y m para que concuerden las unidades de diametro  
    diametro_int = modelos.diametro_interno(Tdiametro)
    material=  modelos.diametro_interno1(Tdiametro)
    Unidades=  modelos.diametro_interno2(Tdiametro)
    Tlargo = convertir(TlargoUni, "m", Tlargo)
    Tlargo2 = convertir(TlargoUni2, "m", Tlargo2)
    longitud = convertir(longitudUni, "mm", longitud)
    #Alto = convertir(AltoUni, "mm", Alto)

    TubeLargo = Tlargo + Tlargo2

    #Separar los componentes de la ubicacion
    lati = float(ubicacion.split(",")[0])
    longi = float(ubicacion.split(",")[1])

    #Convertir a bar para que concuerden las unidades de presion
    presionTub = convertir(presionUni, "bar", presionTub)

    #TODO traer presion atmos para obtener presion total
    presionAtmos = modelos.presion_atmos(elevacion(lati, longi))

    #Se calcula la duracion de la fuga
    tiempoInicio = datetime.datetime(int(tiempoInicio[0:4]), int(tiempoInicio[5:7]), int(tiempoInicio[8:10]), int(tiempoInicio[11:13]), int(tiempoInicio[14:16]), 0, 0)
    tiempoFin = datetime.datetime(int(tiempoFin[0:4]), int(tiempoFin[5:7]), int(tiempoFin[8:10]), int(tiempoFin[11:13]), int(tiempoFin[14:16]), 0, 0)
    duracion = tiempoFin - tiempoInicio
    duracion = duracion.total_seconds()
    horas = int(duracion // 3600)
    horasQ = duracion % 3600
    minutos = int(horasQ // 60)
    duracion2 = duracion / 3600

    #Si la ruptura es total se calculan los valores de una ruptura circular con el diametro de la tubería
    if forma == "total":
        Fuga_diame = diametro_int
        area = modelos.calc_area("circ", Fuga_diame, 0, 0, 0)
        perimetro = modelos.calc_peri("circ", Fuga_diame, 0, 0, 0)
    #Si la ruptura es recta se calcula el diametro hidraulico
    elif forma == "recta":
        if equi == 'on':
            area = modelos.calc_area("circ", Fuga_diame, 0, 0, 0)
            perimetro = modelos.calc_peri("circ", Fuga_diame, 0, 0, 0)
        else:
            area = modelos.calc_area("recta", 0, 0, 0, longitud)
            perimetro = modelos.calc_peri("recta", 0, 0, 0, longitud)
            Fuga_diame = modelos.diametro_hidraulico(area, perimetro,diametro_int)
    else:
        area = modelos.calc_area(forma, Fuga_diame, 0, 0, longitud)
        perimetro = modelos.calc_peri(forma, Fuga_diame, 0, 0, longitud)

    #Valores necesarios para el modelo utp
    #diametro_hidraulico = modelos.diametro_hidraulico(area, presionTub)
    #rho_int = modelos.rho_interior(presionTub, 1, temperatura)

    if forma == "rect" or forma == "recta":
        coef_flujo = 0.9
    elif forma == "tria":
        coef_flujo = 0.95
    else:
        coef_flujo = 1

    if forma == "recta":
        forma = "Recta"
        medida = longitud
        medidaUni = longitudUni
        #medida2=Alto
        #medidaUni2 = AltoUni
    elif forma == "total":
        forma = "Total"
        medida = ""
        medidaUni = ""
        #medida2= ""
        #medidaUni2 = ""
    else:
        forma = "Circular"
        medida = Fdiametro
        medidaUni = FdiametroUni
        #medida2= ""
        #medidaUni2 = ""

    

    

    ####
    # NUEVO METODO DE CALCULO DEL FLUJO
    ####
    R_vals = np.array([0.0, 0.25, 0.5, 0.75, 1.0])
    R_real = 1.0 if forma == "Total" else Fuga_diame / diametro_int
    Q_vals = []
    if diametro_int > 76.2:
        d1 = 50.8
        d2 = 76.2
        Q1_vals = []
        Q2_vals = []
        for R in R_vals:
            R_actual = 1.0 if forma == "Total" else R
            Q1_iter = []
            Q2_iter = []

            for d_tube_i, Q_iter in zip([d1, d2], [Q1_iter, Q2_iter]):
                L0 = modelos.obtener_L0(R_actual, material)
                if TubeLargo <= L0:
                    Q0 = modelos.modelo_utpSuper(Fuga_diame, d_tube_i, presionTub, presionAtmos, subte, direccion, forma,TubeLargo,material,R_actual)
                    Q_iter.append(Q0)
                else:
                    Q0 = modelos.modelo_utpSuper(Fuga_diame, d_tube_i, presionTub, presionAtmos, subte, direccion, forma,L0,material,R_actual)
                    Q_iter.append(Q0)
                    for L in range(L0 + 1, int(TubeLargo) + 1):
                        a = modelos.alpha(L, R_actual, material)
                        Qi = Q_iter[-1] if a is None else Q_iter[-1] * (1 - a)
                        Q_iter.append(Qi)

            Q1_vals.append(Q1_iter[-1])
            Q2_vals.append(Q2_iter[-1])

        Q1_interp = np.interp(R_real, R_vals, Q1_vals)
        Q2_interp = np.interp(R_real, R_vals, Q2_vals)
        Q_extrap = Q1_interp + (Q2_interp - Q1_interp) * ((diametro_int - d1) / (d2 - d1))

        flujo=Q_extrap
    else:
        for R in R_vals:
            R_actual = 1.0 if forma == "Total" else R
            Q_iter = []
            L0 = modelos.obtener_L0(R_actual, material)

            if TubeLargo <= L0:
                Q0 = modelos.modelo_utpSuper(Fuga_diame, diametro_int, presionTub, presionAtmos, subte, direccion, forma,TubeLargo,material,R_actual)
                Q_iter.append(Q0)
            else:
                Q0 = modelos.modelo_utpSuper(Fuga_diame, diametro_int, presionTub, presionAtmos, subte, direccion, forma,L0,material,R_actual)
                Q_iter.append(Q0)
                print(L0,TubeLargo)
                for L in range(L0 + 1, int(TubeLargo) + 1):
                    a = modelos.alpha(L, R_actual, material)
                    Qi = Q_iter[-1] if a is None else Q_iter[-1] * (1 - a)
                    Q_iter.append(Qi)

            Q_vals.append(Q_iter[-1])

        Q_final = np.interp(R_real, R_vals, Q_vals)
        flujo=Q_final
    #Se calcula el flujo y volumen total perdido

    #flujo = coef_flujo * modelos.modelo_utpSuper(Fuga_diame, diametro_int, presionTub, presionAtmos, subte, direccion, forma)
    vol_muerto = modelos.vol_muerto(diametro_int, TubeLargo)
    Volumenfugado=(flujo * duracion2)
    volumen = Volumenfugado + vol_muerto
    TubeLargo=convertir("m", TlargoUni, TubeLargo)

    #Formatear valor mostrado en caso de que sea muy pequeño
    if flujo < 1:
        if round(flujo, 2) == 0:
            flujo2 = "{:.2e}".format(flujo)
        else:
            flujo2 = round(flujo, 2)
    else:
        flujo2 = round(flujo, 2)
    if volumen < 1:
        if round(volumen, 2) == 0:
            volumen2 = "{:.2e}".format(volumen)
        else:
            volumen2 = round(volumen, 2)
    else:
        volumen2 = round(volumen, 2)
    if Volumenfugado < 1:
        if round(Volumenfugado, 2) == 0:
            volumenfugado2 = "{:.2e}".format(Volumenfugado)
        else:
            volumenfugado2 = round(Volumenfugado, 2)
    else:
        volumenfugado2 = round(Volumenfugado, 2)
    if vol_muerto < 1:
        if round(vol_muerto, 2) == 0:
            vol_muerto2 = "{:.2e}".format(vol_muerto)
        else:
            vol_muerto2 = round(vol_muerto, 2)
    else:
        vol_muerto2 = round(vol_muerto, 2)
    
    hoy = datetime.datetime.now()
    vol_muerto2 = float(vol_muerto2)
    resp = make_response(render_template('resultados.html',Unidades=Unidades,material=material,Tdiametro=round(Tdiametro,2), Volumenfugado=round(Volumenfugado,2),vol_muerto=round(vol_muerto2,2),Subte=subte,Tlargo=TubeLargo,TlargoUni=TlargoUni, orden=orden, area=round(area,2), flujo=round(flujo,2), volumen=round(volumen,2), horas=horas, minutos=minutos, longitud=round(longi,4), latitud=lati, año_reg = hoy.year, mes_reg = hoy.month, dia_reg = hoy.day, año_inicio=tiempoInicio.year, mes_inicio=tiempoInicio.month, dia_inicio=tiempoInicio.day, direccion = "Unidireccional" if direccion == "uni" else "Bidireccional", presion_tube = round(convertir("bar", "psig", presionTub),2), presion_atmos=round(convertir("bar", "psig",presionAtmos),2), forma=forma))
    if request.cookies.get('guardado') == "false":
        result = guardar_evento(orden, ubicacion, convertir("bar", "psig", presionTub), subte, Tlargo, TlargoUni, Tlargo2, TlargoUni2, Tdiametro, material, Unidades, direccion, forma, medida, medidaUni, area, flujo, volumen, vol_muerto, Volumenfugado, tiempoInicio, duracion, hoy, presionAtmos, escape if equi == "on" else 'no')
        if (result):
            resp.set_cookie('guardado', 'true')
            return resp
        else:
            resp = make_response(redirect("/NuevoEvento", code=307))
            resp.set_cookie('nuevo', '-1')
            return resp
        
    

@app.route('/CargarBuscar', methods=['POST', 'GET'])

def cargarBuscar():
    resp = make_response(redirect('/BuscarEvento'))
    resp.set_cookie('orden', '')
    return resp

@app.route('/BuscarEvento', methods=['POST', 'GET'])

def renderBuscar():
    resp = make_response(render_template('buscar.html'))
    return resp

@app.route('/Buscar', methods=['POST'])
def buscar():
    orden = request.form.get('orden')
    ## OBTENEMOS LA ORDEN EN FORMATO STRING 
    objeto_ = {'orden':orden }
    
    url = dominio+"rupture/getSpecificEvent"#"https://mongorupture.efigasprojecthub.site/rupture/login"
    
    print("DATOS QUERY: ",url,objeto_)
    response = requests.post(url,json=objeto_)
    if(response.json()['status'] == 'Si hay elemento'):
        resp = make_response(redirect('/Reporte', code=307))
        resp.set_cookie('orden', orden)
        return resp 
    else:
        resp = make_response(redirect('/BuscarEvento'))
        resp.set_cookie('orden', '-1')
        return resp
    
    



@app.route('/Reporte', methods=['POST'])
def reporte():
    orden = request.cookies.get("orden")

    ## OBTENEMOS LA ORDEN EN FORMATO STRING 
    objeto_ = {'orden':orden }
    
    url = dominio+"rupture/getSpecificEvent"#"https://mongorupture.efigasprojecthub.site/rupture/login"

    response = requests.post(url,json=objeto_)
    if(response.json()['status'] == 'Si hay elemento'):
        fila = response.json()['info']
        ubicacion = fila["ubicacion"]
        presion = float(fila["presion"])
        subte = fila["subte"]
        dist_tube = float(fila["dist_tube"])
        dist_tube_uni = fila["dist_tube_uni"]
        dist_tube2 = float(fila["dist_tube2"])
        dist_tube_uni2 = fila["dist_tube_uni2"]
        diame_tube = float(fila["diame_tube"])
        Unidades = fila["Unidades"]
        material = fila["Material"]
        direccion = fila["direccion"]
        forma = fila["forma"]
        area = float(fila["area"])
        flujo = float(fila["flujo"])
        volumen = float(fila["volumen"])
        aprobado = fila["aprobado"]
        volumen_muerto = float(fila["volumen_muerto"])
        volumen_fuga = float(fila["volumen_fuga"])
        inicio = pd.Timestamp( datetime.datetime.strptime(fila["inicio"], '%Y-%m-%d %H:%M')).to_pydatetime()
        duracion = fila["duracion"]
        hora_reg = pd.Timestamp( datetime.datetime.strptime(fila["hora_reg"], '%Y-%m-%d %H:%M')).to_pydatetime()
        presion_atmos = float(fila["presion_atmos"])
        duracion = float(duracion)
        horas = int(duracion // 3600)
        horasQ = duracion % 3600
        minutos = int(horasQ // 60)
        #Separar los componentes de la ubicacion
        lati = float(ubicacion.split(",")[0])
        longi = float(ubicacion.split(",")[1])

        largo2 = convertir(dist_tube_uni2, dist_tube_uni, dist_tube2)
        largo = dist_tube + largo2

        if flujo < 1:
            if round(flujo, 2) == 0:
                flujo2 = "{:.2e}".format(flujo)
            else:
                flujo2 = round(flujo, 2)
        else:
            flujo2 = round(flujo, 2)
        if volumen < 1:
            if round(volumen, 2) == 0:
                volumen2 = "{:.2e}".format(volumen)
            else:
                volumen2 = round(volumen, 2)
        else:
            volumen2 = round(volumen, 2)
        if volumen_fuga < 1:
            if round(volumen_fuga, 2) == 0:
                volumen_fuga2 = "{:.2e}".format(volumen_fuga)
            else:
                volumen_fuga2 = round(volumen_fuga, 2)
        else:
            volumen_fuga2 = round(volumen_fuga, 2)
        if volumen_muerto < 1:
            if round(volumen_muerto, 2) == 0:
                vol_muerto2 = "{:.2e}".format(volumen_muerto)
            else:
                vol_muerto2 = round(volumen_muerto, 2)
        else:
            vol_muerto2 = round(volumen_muerto, 2)

        resp = make_response(render_template('reporte.html',material=material, Unidades=Unidades,diame_tube=round(diame_tube,2), orden = orden, Volumenfugado=volumen_fuga2, vol_muerto=vol_muerto2,Subte=subte,Tlargo=largo,TlargoUni=dist_tube_uni, area=round(area,2), flujo=flujo2, volumen=volumen2, horas=horas, minutos=minutos, longitud=longi, latitud=lati, año_reg = hora_reg.year, mes_reg = hora_reg.month, dia_reg = hora_reg.day, año_inicio=inicio.year, mes_inicio=inicio.month, dia_inicio=inicio.day, direccion = "Unidireccional" if direccion == "uni" else "Bidireccional", presion_tube = round(presion,2), presion_atmos=round(convertir("bar", "psig",presion_atmos),2), forma=forma, aprobado = aprobado))
        return resp
    else:
        resp = make_response(redirect('/BuscarEvento'))
        resp.set_cookie('orden', '-1')
        return resp


@app.route('/Editar', methods=['POST'])
def editar():
    orden = request.cookies.get("orden")
    ## OBTENEMOS LA ORDEN EN FORMATO STRING 
    objeto_ = {'orden':orden }
    
    url = dominio+"rupture/getSpecificEvent"#"https://mongorupture.efigasprojecthub.site/rupture/login"
    
    response = requests.post(url,json=objeto_)
    if(response.json()['status'] == 'Si hay elemento'):
            fila = response.json()['info']
            ubicacion = fila["ubicacion"]
            presion = fila["presion"]
            subte = fila["subte"]
            dist_tube = fila["dist_tube"]
            dist_tube_uni = fila["dist_tube_uni"]
            dist_tube2 = fila["dist_tube2"] if fila["dist_tube2"] > 0 else ""
            dist_tube_uni2 = fila["dist_tube_uni2"]
            diame_tube = fila["diame_tube"]
            flujo = fila["direccion"]
            forma = fila["forma"]
            medida_rupt = fila["medida_rupt"]
            medida_uni = fila["medida_uni"]
            inicio = pd.Timestamp(datetime.datetime.strptime(fila["inicio"], '%Y-%m-%d %H:%M')).to_pydatetime()
            duracion = int(fila["duracion"])
            diame_equi = fila["diame_equi"]
            diame_tube = float_a_int(float(diame_tube))
            delta = datetime.timedelta(seconds=duracion)
            fin = inicio + delta
            fecha_inicio = inicio.strftime('%Y-%m-%dT%H:%M')
            fecha_fin = fin.strftime('%Y-%m-%dT%H:%M')

            resp = make_response(render_template("editar.html", orden=orden, ubicacion=ubicacion, tiempoInicio=fecha_inicio, tiempoFin=fecha_fin, presion=round(presion,0), DiameTube=diame_tube, Flujo=flujo, DistTube=dist_tube, DistTubeUni=dist_tube_uni, DistTube2=dist_tube2, DistTubeUni2=dist_tube_uni2, subte=subte, Forma=forma, diameEqui=diame_equi,medida_rupt=medida_rupt,medida_uni=medida_uni))
    else:
        resp = make_response("Error al abrir el archivo")
    
    return resp

@app.route('/Editado', methods=['POST'])
def editarEvento():
    # Se obtienen todas las entradas
    
    orden = request.form.get('orden')
    ubicacion = request.form.get('ubicacion')
    presionTub = request.form.get('presion')
    presionUni = request.form.get('presionUni')
    subte = request.form.get('subte')
    equi = request.form.get('diameEqui')
    escape = request.form.get('escape')
    direccion = request.form.get("Flujo")
    forma = request.form.get('Forma')
    Fdiametro = request.form.get('DiameFuga')
    longitud = request.form.get('LongiFuga')
    Alto = request.form.get('Altofuga')
    FdiametroUni = request.form.get('DiameFugaUni')
    longitudUni = request.form.get('LongiFugaUni')
    AltoUni = request.form.get('LongiFugaUni')
    Tlargo = request.form.get("DistTube")
    TlargoUni = request.form.get("DistTubeUni")
    Tlargo2 = request.form.get("DistTube2")
    TlargoUni2 = request.form.get("DistTubeUni2")
    Tdiametro = request.form.get('DiameTube')
    tiempoInicio = request.form.get('tiempoInicio')
    tiempoFin = request.form.get('tiempoFin')

    #Se castean los valores porque a numpy no le gusta usar cadenas
    presionTub = float(presionTub)
    Fdiametro = float(Fdiametro) if Fdiametro != "" else 0
    Tlargo = float(Tlargo) if Tlargo != "" else 0
    Tlargo2 = float(Tlargo2) if Tlargo2 != "" else 0
    longitud = float(longitud) if longitud != "" else 0
    Alto = float(Alto) if Alto != "" else 0
    Tdiametro = float(Tdiametro)
    
    #Aplicar los diametros equivalentes
    if equi == 'on':
        Fuga_diame = convertir("in", "mm", modelos.diametro_equi(Tdiametro, escape))
    else:
        Fuga_diame = convertir(FdiametroUni, "mm", Fdiametro)
    #Convertir a mm y m para que concuerden las unidades de diametro  
    diametro_int = modelos.diametro_interno(Tdiametro)
    material=  modelos.diametro_interno1(Tdiametro)
    Unidades=  modelos.diametro_interno2(Tdiametro)
    Tlargo = convertir(TlargoUni, "m", Tlargo)
    Tlargo2 = convertir(TlargoUni2, "m", Tlargo2)
    longitud = convertir(longitudUni, "mm", longitud)
    #Alto = convertir(AltoUni, "mm", Alto)

    TubeLargo = Tlargo + Tlargo2

    #Separar los componentes de la ubicacion
    lati = float(ubicacion.split(",")[0])
    longi = float(ubicacion.split(",")[1])

    #Convertir a bar para que concuerden las unidades de presion
    presionTub = convertir(presionUni, "bar", presionTub)

    #TODO traer presion atmos para obtener presion total
    presionAtmos = modelos.presion_atmos(elevacion(lati, longi))

    #Se calcula la duracion de la fuga
    tiempoInicio = datetime.datetime(int(tiempoInicio[0:4]), int(tiempoInicio[5:7]), int(tiempoInicio[8:10]), int(tiempoInicio[11:13]), int(tiempoInicio[14:16]), 0, 0)
    tiempoFin = datetime.datetime(int(tiempoFin[0:4]), int(tiempoFin[5:7]), int(tiempoFin[8:10]), int(tiempoFin[11:13]), int(tiempoFin[14:16]), 0, 0)
    duracion = tiempoFin - tiempoInicio
    duracion = duracion.total_seconds()
    horas = int(duracion // 3600)
    horasQ = duracion % 3600
    minutos = int(horasQ // 60)
    duracion2 = duracion / 3600

    #Si la ruptura es total se calculan los valores de una ruptura circular con el diametro de la tubería
    if forma == "total":
        Fuga_diame = diametro_int
        area = modelos.calc_area("circ", Fuga_diame, 0, 0, 0)
        perimetro = modelos.calc_peri("circ", Fuga_diame, 0, 0, 0)
    #Si la ruptura es recta se calcula el diametro hidraulico
    elif forma == "recta":
        if equi == 'on':
            area = modelos.calc_area("circ", Fuga_diame, 0, 0, 0)
            perimetro = modelos.calc_peri("circ", Fuga_diame, 0, 0, 0)
        else:
            area = modelos.calc_area("recta", 0, 0, 0, longitud)
            perimetro = modelos.calc_peri("recta", 0, 0, 0, longitud)
            Fuga_diame = modelos.diametro_hidraulico(area, perimetro,diametro_int)
    else:
        area = modelos.calc_area(forma, Fuga_diame, 0, 0, longitud)
        perimetro = modelos.calc_peri(forma, Fuga_diame, 0, 0, longitud)

    #Valores necesarios para el modelo utp
    #diametro_hidraulico = modelos.diametro_hidraulico(area, presionTub)
    #rho_int = modelos.rho_interior(presionTub, 1, temperatura)

    if forma == "rect" or forma == "recta":
        coef_flujo = 0.9
    elif forma == "tria":
        coef_flujo = 0.95
    else:
        coef_flujo = 1

    if forma == "recta":
        forma = "Recta"
        medida = longitud
        medidaUni = longitudUni
        #medida2 = Alto
        #medidaUni2 = AltoUni
    elif forma == "total":
        forma = "Total"
        medida = ""
        medidaUni = ""
        #medida2=""
        #medidaUni2=""
    else:
        forma = "Circular"
        medida = Fdiametro
        medidaUni = FdiametroUni
        #medida2=""
        #medidaUni2=""
    
    #diametro_int de mm a m
    # Fuga_diame /= 1000
    # diametro_int /= 1000
    #Se calcula el flujo y volumen total perdido
    flujo = coef_flujo * modelos.modelo_utpSuper(Fuga_diame, diametro_int, presionTub, presionAtmos, subte, direccion, forma)
    vol_muerto = modelos.vol_muerto(diametro_int, TubeLargo)
    Volumenfugado=(flujo * duracion2)
    volumen = Volumenfugado + vol_muerto
    TubeLargo=convertir("m", TlargoUni, TubeLargo)
    
    resp = make_response(redirect("/Reporte", code=307))
    if request.cookies.get('guardado') == "false":
        response = editar_evento(orden, ubicacion, convertir("bar", "psig", presionTub), subte, Tlargo, TlargoUni, Tlargo2, TlargoUni2, Tdiametro, material, Unidades, direccion, forma, medida, medidaUni,area, flujo, volumen, vol_muerto, Volumenfugado, tiempoInicio, duracion, presionAtmos, escape if equi == "on" else 'no')
        print("EDITADO? ",response)
        resp.set_cookie('guardado', 'true')
    return resp

@app.route('/Aprobar', methods=['POST'])
def aprobar():
    archivo = "eventos/" + request.cookies.get("empresa") + ".xlsx"
    workbook = load_workbook(archivo)
    sheet = workbook.worksheets[0]
    orden = request.cookies.get("orden")
    for row in sheet.iter_rows():  # Iterando en la primera columna
        if row[0].value == orden:
            row[26].value = "sí"
            break
    workbook.save(archivo)
    workbook.close()
    return make_response(redirect('/Reporte', code=307))

@app.route('/Descargar', methods=['POST'])
def downloadFile ():
    
    ### Llamamos al endpoint para traernos la lista de eventos 
    # url = dominio+"rupture/getSpecificEvent"#"https://mongorupture.efigasprojecthub.site/rupture/login"
    url = dominio+"rupture/getEvents"
    response = requests.get(url) 
    # Crear un DataFrame de pandas a partir de la lista de objetos
    df = pd.DataFrame(response.json())
     # Crear un buffer en memoria para guardar el archivo Excel
    buffer = BytesIO()

    # Guardar el DataFrame en el buffer como archivo Excel
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Datos')

    # Regresar el cursor del buffer al inicio para que Flask pueda leerlo
    buffer.seek(0)

    # Enviar el archivo como una respuesta de descarga
    return send_file(
        buffer,
        as_attachment=True,
        download_name='datos.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    # response = send_file(archivo, as_attachment=True)
    # return send_file(archivo, as_attachment=True)

#Carga el archivo en cache para poder borrarlo sin tener problemas al enviar
@app.route('/Descargar/<path:filename>', methods=['POST'])
def download(filename):
    path = os.path.join(
        app.root_path,
        'temp',
        filename
    )
    cache = io.BytesIO()
    with open(path, 'rb') as fp:
        shutil.copyfileobj(fp, cache)
        cache.flush()
    cache.seek(0)
    os.remove(path)
    return send_file(cache, as_attachment=True, download_name=filename)

def convertir (origen, objetivo, valor):
    conv = valor
    #Distancia
    if origen == "mm":
        if objetivo == "m":
            conv = valor/1000
        elif objetivo == "in":
            conv = valor/25.4
        elif objetivo == "ft":
            conv = valor/304.8
        elif objetivo == "mm":
            conv = valor
    elif origen == "ft":
        if objetivo == "m":
            conv = valor/3.281
        elif objetivo == "in":
            conv = valor/12
        elif objetivo == "mm":
            conv = valor*304.8
        elif objetivo == "ft":
            conv = valor
    elif origen == "in":
        if objetivo == "mm":
            conv = valor*25.4
        elif objetivo == "m":
            conv = valor/39.37
        elif objetivo == "ft":
            conv = valor*12
        elif objetivo == "in":
            conv = valor
    elif origen == "m":
        if objetivo == "mm":
            conv = valor/1000
        elif objetivo == "in":
            conv = valor*39.37
        elif objetivo == "ft":
            conv = valor*3.281
        elif objetivo == "m":
            conv = valor
    #Presion
    elif origen == "psig":
        if objetivo == "bar":
            conv = valor/14.504
        elif objetivo == "pascal":
            conv = valor/6894.76
        elif objetivo == "psig":
            conv = valor
    elif origen == "bar":
        if objetivo == "psig":
            conv = valor*14.504
        elif objetivo == "pascal":
            conv = valor*100000
        elif objetivo == "bar":
            conv = valor
    elif origen == "pascal":
        if objetivo == "psig":
            conv = valor*6894.76
        elif objetivo == "bar":
            conv = valor/100000
        elif objetivo == "pascal":
            conv = valor
    #Temperatura
    if origen == "c":
        if objetivo == "f":
            conv = (valor*9/5) + 32
    return conv

def guardar_evento(orden, ubicacion, presion_tube, subte, dist_tube, dist_tube_uni, dist_tube2, dist_tube_uni2, diame_tube, material, Unidades, dire, forma, medida_fuga, medida_uni,area, flujo, volumen, volumen_muerto, volumen_fuga, inicio, duracion, fecha, presion_atmos, diame_equi):
    #EN CASO DE EDITAR LAS COLUMNAS QUE SE GUARDAN
    #Se debe editar también la función de creación de tabla y función de aprobación para que
    #el index concuerde con la columna de aprobado
    
    objeto_ = {'orden':orden, 'ubicacion':ubicacion,'presion':presion_tube, 'subte':subte, 'dist_tube':dist_tube, 'dist_tube_uni':dist_tube_uni, 'dist_tube2':dist_tube2, 'dist_tube_uni2':dist_tube_uni2,'diame_tube':diame_tube,'Material':material, 'Unidades':Unidades , 'direccion':dire, 'forma':forma, 'medida_rupt':medida_fuga, 'medida_uni':medida_uni,'area':area, 'flujo':float(flujo), 'volumen':float(volumen), 'inicio':inicio.strftime('%Y-%m-%d %H:%M'), 'duracion':duracion, 'hora_reg':fecha.strftime('%Y-%m-%d %H:%M'), 'presion_atmos':float(presion_atmos), 'volumen_fuga':float(volumen_fuga),'volumen_muerto':float(volumen_muerto), 'diame_equi':diame_equi, 'aprobado':'no' }
    
    url = dominio+"rupture/createEvent"#"https://mongorupture.efigasprojecthub.site/rupture/login"
    response = requests.post(url,json=objeto_)
    if(response.json()['status'] == 'Orden creada con éxito'):
        return True 
    else:
        return False

def editar_evento(orden, ubicacion, presion_tube, subte, dist_tube, dist_tube_uni, dist_tube2, dist_tube_uni2, diame_tube, material, Unidades, dire, forma, medida_fuga, medida_uni, area, flujo, volumen, volumen_muerto, volumen_fuga, inicio, duracion, presion_atmos, diame_equi):
    objeto_ = {'orden':orden, 'ubicacion':ubicacion,'presion':presion_tube, 'subte':subte, 'dist_tube':dist_tube, 'dist_tube_uni':dist_tube_uni, 'dist_tube2':dist_tube2, 'dist_tube_uni2':dist_tube_uni2,'diame_tube':diame_tube,'Material':material, 'Unidades':Unidades , 'direccion':dire, 'forma':forma, 'medida_rupt':medida_fuga, 'medida_uni':medida_uni, 'area':area, 'flujo':float(flujo), 'volumen':float(volumen), 'inicio':inicio.strftime('%Y-%m-%d %H:%M'), 'duracion':duracion, 'presion_atmos':float(presion_atmos), 'volumen_fuga':float(volumen_fuga),'volumen_muerto':float(volumen_muerto), 'diame_equi':diame_equi, 'aprobado':'no' }
    url = dominio+"rupture/updateEvent"#"https://mongorupture.efigasprojecthub.site/rupture/login"
    
    response = requests.post(url,json=objeto_)
    
    if(response.json()['success']):
        return True 
    else:
        return False


def guardar_token(email, token, tipo):
    hoy = datetime.datetime.now()
    wb = load_workbook("solicitudes.xlsx")
    ws = wb.active
    for row in ws.iter_rows():
        if row[0].value == email and row[1].value == tipo:
            row[2].value = hoy
            row[3].value = token 
            wb.save("solicitudes.xlsx")
            wb.close()
            return
    ws.append([email, tipo, hoy, token])
    wb.save("solicitudes.xlsx")
    wb.close()
    return

def unico(orden):
    ### OBTENER SI EL NUMERO DE ORDEN EXISTE 
    return True

def elevacion(latitud, longitud):
    valor = 0
    api = "https://api.open-elevation.com/api/v1/lookup?locations=" + str(latitud) + "," + str(longitud)
    try:
        #Respuesta es un json que se decodifica como dict
        respuesta = requests.get(api).json()
        #Dentro del dict hay una llave results con array de un elemento con otro dict adentro
        valor = respuesta.get('results')[0].get('elevation')
    except:
        valor = 0
    return valor

def float_a_int(valor):
    if valor.is_integer():
        return int(valor)
    else:
        return valor
    
def crear_tabla(empresa):
    wb = Workbook()
    ws = wb.active
    headers = ['orden', 'ubicacion', 'presion', 'subte', 'dist_tube', 'dist_tube_uni', 'dist_tube2', 'dist_tube_uni2', 'diame_tube', 'Material', 'Unidades', 'direccion', 'forma', 'medida_rupt', 'medida_uni','area', 'flujo', 'volumen', 'inicio', 'duracion', 'presion_atmos', 'volumen_fuga', 'volumen_muerto', 'diame_equi', 'aprobado', 'hora_reg']
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)
    wb.save("eventos/" + empresa + ".xlsx")
    wb.close()

if __name__=='__main__':
    app.run(debug=True,port=5000) ### SI ES NECESARIO DESDE AQUI SE CAMBIA EL PUERTO DE EJECUCIÓN DE LA APLICACIÓN Y ESTA ASOCIADO TAMBIEN A LA VARIABLE DOMINIO POR ENDE SI 
    ### SE CAMBIA SU PUERTO TENER EN CUANTO EN EL DOMINIO.
    ### LOS ARCHIVOS DE .xlsx ESTAN SIMULANDO UNA CONSULTA Y SOBRE ESCRITURA A UNA BASE DE DATOS, PERO EN CASO DE PENSAR EN UN DESPLIEGUE SE HACE NECESARIO MODIFICAR EL CODIGO PRESENTE PARA 
    ### REALIZAR CONSULTAS Y ASOCIAR TODO A UNA BASE DE DATOS ESPECIFICA.